#Create an Excel with data of Student database and add all the values which is required for student management database, Read the excel file and add the data into a DB (using script)

import openpyxl
path = "students.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row = 5, column = 3)
print(cell_obj.value)

for i in range(1,11):
    cell_obj = sheet_obj.cell(row = 5, column = i)
    print(cell_obj.value)
import mysql.connector

mydb = mysql.connector.connect(
  host="localhost",
 user="root",
  password="Singapore@22",
)

mycursor = mydb.cursor()
print(mydb)
dbse = mydb.cursor()

dbse.execute("CREATE DATABASE Students_Management_System")
dbse = mydb.cursor()

dbse.execute("SHOW DATABASES")

for entry in dbse:
  print(entry)

mydb = mysql.connector.connect(
  host="localhost",
 user="root",
  password="Singapore@22",
  database="students_management_system"
)
dbse = mydb.cursor()

dbse.execute("CREATE TABLE student (Reg_no INT(10),Name VARCHAR(255), Semester1_result INT(10),Semester2_result INT(10),Semester3_result INT(10),Semester4_result INT(10),Total_CGPA INT(35) ,Mobile_Number INT(10),Email_id VARCHAR(55))")
dbse = mydb.cursor()

dbse.execute("SHOW TABLES")

for value in dbse:
  print(value)
cur = mydb.cursor()
cur.execute('SELECT * FROM student')
for row in cur:
    print(row)
import pandas as pd

df = pd.read_excel('students.xlsx')
import xlrd
import MySQLdb
xl_sheet = xlrd.open_workbook("students.xlsx")
xl_sheet
sheet_name =xl_sheet.sheet_names()
sheet_name
mydb = mysql.connector.connect(
  host="localhost",
 user="root",
  password="Singapore@22",
  database="students_management_system"
)

cur = mydb.cursor()
for s in range(0,1):
    sheet=xl_sheet.sheet_by_index(s)
    sql= "INSERT INTO student(Reg_no,Name,Semester1_result,Semester2_result ,Semester3_result, Total_CGPA ,Mobile_Number Email_id) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)"
    for r in range(1,sheet.nrows):
        Reg_no =sheet.cell(r,0).value
        Name =sheet.cell(r,1).value
        Semester1_result =sheet.cell(r,2).value
        Semester2_result =sheet.cell(r,3).value
        Semester3_result =sheet.cell(r,4).value
        Total_CGPA =sheet.cell(r,5).value
        Mobile_Number =sheet.cell(r,6).value
        Email_id=sheet.cell(r,7).value
        values =(Reg_no,Name ,Semester1_result,Semester2_result ,Semester3_result, Total_CGPA ,Mobile_Number,Email_id)
        
        cur.execute(sql,values)
mydb.commit()
mycursor = mydb.cursor()

mycursor.execute("SELECT * FROM student")

myresult = mycursor.fetchall()

for x in myresult:
    print(x)
mycursor = mydb.cursor()

mycursor.execute("SELECT NAME FROM student WHERE CGPA >6")

myresult = mycursor.fetchall()

for x in myresult:
    print(x)
mydb.commit()
mydb.close()